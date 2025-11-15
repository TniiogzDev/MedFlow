# |========================================================================|
# |                           IMPORTACIONES                                |
# |========================================================================|
from django.shortcuts import render                                           # Para renderizar plantillas HTML
from django.shortcuts import render, redirect                                 # Para redireccionar a otras vistas
from django.contrib.auth import authenticate, login, logout                   # Para manejar la autenticación de usuarios
from django.contrib import messages                                           # Para mostrar mensajes al usuario
from django.contrib.auth.decorators import login_required                     # Para proteger vistas que requieren autenticación
from django.http import JsonResponse, HttpResponse, Http404                   # Para manejar respuestas HTTP  
from django.urls import reverse                                               # Para obtener URLs a partir de nombres de vistas  
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie      # Para manejar protección CSRF
from django.core.serializers import serialize                                 # Para serializar datos
from django.utils import timezone                                             # Para manejar zonas horarias    
from django.conf import settings                                              # Para acceder a configuraciones del proyecto
from django.contrib.auth.models import Group                            
from django.contrib.auth import get_user_model                          
from django.contrib.admin.models import LogEntry                              # Para registrar acciones en el admin
from django.db.models import Q                                                # Para consultas complejas en la base de datos
import json                                                                   # Para manejar datos en formato JSON   
from .models import FormularioAtencion, MensajeChat                           # Modelos específicos de la aplicación    
import io                                                                     # Para manejar flujos de datos en memoria  
from openpyxl import Workbook                                                 # Para crear y manipular archivos Excel 
from openpyxl.styles import Font, Alignment, PatternFill                      # Para dar estilo a celdas en Excel
from reportlab.pdfgen import canvas                                           # Para generar archivos PDF 
from reportlab.lib.pagesizes import letter                                    # Para definir tamaños de página en PDF
from reportlab.lib.units import inch                                          # Para manejar unidades de medida en PDF  

# ================== -->
# INICIO CAMBIOS 4.1 -->
# (Importar textwrap para la corrección del PDF)
# ================== -->
import textwrap
# ================== -->
# FIN CAMBIOS 4.1    -->
# ================== -->

# |========================================================================|
# |                   VISTAS DE AUTENTICACIÓN Y MENÚS                      |
# |========================================================================|

# Vista de login
def login_view(request):
    if request.method == 'POST':
        usuario = request.POST.get('username')
        clave = request.POST.get('password')
        user = authenticate(request, username=usuario, password=clave)
        if user is not None:
            login(request, user)
            # Redirección normal (Supervisor va a su propio menú)
            if user.is_superuser or user.groups.filter(name='Supervisor').exists():
                return redirect('menu_supervisor') 
            elif user.groups.filter(name='Ambulancia').exists():
                return redirect('menu_ambulancia')
            elif user.groups.filter(name='Recepcion').exists():
                return redirect('menu_recepcion')
            else:
                # --- MODIFICACIÓN SWEETALERT ---
                context = {
                    'no_role_error': "Parece que aún no estás designado a un rol, espera a que se te asigne uno para continuar."
                }
                return render(request, 'index.html', context)
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
            return render(request, 'index.html', {'error': 'Usuario o contraseña incorrectos.'})
    return render(request, 'index.html')

# Vista de logout
def logout_view(request):
    logout(request)
    return redirect('login')

# Vistas de menús según rol
@login_required 
def menu_ambulancia(request):
    return render(request, 'menu_ambulancia.html')

# Vista de menú para Recepción
@login_required 
@ensure_csrf_cookie 
def menu_recepcion(request):
    return render(request, 'menu_recepcion.html')

# Vista de menú para Supervisor
@login_required
def menu_supervisor(request):
    User = get_user_model() 
    try:
        usuarios_para_auditar = User.objects.exclude(id=request.user.id).prefetch_related('groups')
    except Exception as e:
        print(f"Error al obtener usuarios: {e}")
        usuarios_para_auditar = [] 
    context = { 'usuarios_para_auditar': usuarios_para_auditar }
    return render(request, 'menu_supervisor.html', context)

# |========================================================================|
# |                 VISTAS DE FORMULARIOS (Ambulancia)                     |
# |========================================================================|

# Vista del formulario de ambulancia
@login_required
def formulario_ambulancia(request):
    return render(request, 'formulario_ambulancia.html')

# Vista para enviar el formulario de ambulancia
@login_required 
def enviar_formulario_ambulancia(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            nuevo_formulario = FormularioAtencion(
                creado_por=request.user,
                
                # --- Campos Paso 1 ---
                nombre_paciente=data.get('nombre'),
                rut_paciente=data.get('rut'),
                edad_paciente=data.get('edad') or None, 
                unidad_edad_paciente=data.get('unidadEdad'),
                sexo_paciente=data.get('sexo'),
                prevision=data.get('prevision'),
                accidente_laboral=data.get('accidenteLaboral', False),
                
                # --- Campos Paso 2 ---
                presion_arterial=data.get('presionArterial'),
                frecuencia_cardiaca=data.get('frecuenciaCardiaca') or None,
                frecuencia_respiratoria=data.get('frecuenciaRespiratoria') or None,
                temperatura=data.get('temperatura') or None,
                saturacion_oxigeno=data.get('saturacionOxigeno') or None,
                glasgow=data.get('glasgow') or None,
                llene_capilar=data.get('lleneCapilar'),
                score_mottling=data.get('scoreMottling'),
                musculatura_accesoria=data.get('musculaturaAccesoria', False),
                fio2=data.get('fio2') or None,
                
                # --- Campos Paso 3 ---
                motivo_consulta=data.get('motivoConsulta'),
                antecedentes=data.get('antecedentes'),
                tratamiento_administrado=data.get('tratamiento'),
                solicitudes_paciente=data.get('solicitudesPaciente'),
                funcionalidad=data.get('funcionalidad'),
                prestacion_requerida=data.get('prestacionRequerida'),

                # --- Campos Paso 4 ---
                triage=data.get('triage'),
                instrucciones_recepcion=data.get('instrucciones'),
                eta_fecha=data.get('etaDate') or None, 
                eta_hora=data.get('etaTime') or None, 
            )
            
            nuevo_formulario.save()
            return JsonResponse({'status': 'exito', 'case_id': nuevo_formulario.id}, status=201)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# Vista para editar un formulario existente
@login_required
def editar_formulario_view(request, form_id):
    try:
        formulario = FormularioAtencion.objects.get(id=form_id)
    except FormularioAtencion.DoesNotExist:
        messages.error(request, 'Formulario no encontrado.')
        return redirect('menu_ambulancia')
    if formulario.creado_por != request.user:
        messages.error(request, 'No tienes permiso para editar este formulario.')
        return redirect('menu_ambulancia')
    context = { 'formulario': formulario }
    return render(request, 'editar_formulario.html', context)

# Vista para actualizar un formulario existente
@login_required
def actualizar_formulario_view(request, form_id):
    if request.method == 'POST':
        try:
            formulario = FormularioAtencion.objects.get(id=form_id)
            
            if formulario.creado_por != request.user:
                return JsonResponse({'status': 'error', 'message': 'No tienes permiso para editar este formulario.'}, status=403)

            data = json.loads(request.body)
            
            # --- Campos Paso 1 ---
            formulario.nombre_paciente = data.get('nombre', formulario.nombre_paciente)
            formulario.rut_paciente = data.get('rut', formulario.rut_paciente)
            formulario.edad_paciente = data.get('edad') or formulario.edad_paciente
            formulario.unidad_edad_paciente = data.get('unidadEdad', formulario.unidad_edad_paciente)
            formulario.sexo_paciente = data.get('sexo', formulario.sexo_paciente)
            formulario.prevision = data.get('prevision', formulario.prevision)
            formulario.accidente_laboral = data.get('accidenteLaboral', formulario.accidente_laboral)

            # --- Campos Paso 2 ---
            formulario.presion_arterial = data.get('presionArterial', formulario.presion_arterial)
            formulario.frecuencia_cardiaca = data.get('frecuenciaCardiaca') or formulario.frecuencia_cardiaca
            formulario.frecuencia_respiratoria = data.get('frecuenciaRespiratoria') or formulario.frecuencia_respiratoria
            formulario.temperatura = data.get('temperatura') or formulario.temperatura
            formulario.saturacion_oxigeno = data.get('saturacionOxigeno') or formulario.saturacion_oxigeno
            formulario.glasgow = data.get('glasgow') or formulario.glasgow
            formulario.llene_capilar = data.get('lleneCapilar', formulario.llene_capilar)
            formulario.score_mottling = data.get('scoreMottling', formulario.score_mottling)
            formulario.musculatura_accesoria = data.get('musculaturaAccesoria', formulario.musculatura_accesoria)
            formulario.fio2 = data.get('fio2') or formulario.fio2
            
            # --- Campos Paso 3 ---
            formulario.motivo_consulta = data.get('motivoConsulta', formulario.motivo_consulta)
            formulario.antecedentes = data.get('antecedentes', formulario.antecedentes)
            formulario.tratamiento_administrado = data.get('tratamiento', formulario.tratamiento_administrado)
            formulario.solicitudes_paciente = data.get('solicitudesPaciente', formulario.solicitudes_paciente)
            formulario.funcionalidad = data.get('funcionalidad', formulario.funcionalidad)
            formulario.prestacion_requerida = data.get('prestacionRequerida', formulario.prestacion_requerida)
            
            # --- Campos Paso 4 ---
            formulario.triage = data.get('triage', formulario.triage)
            formulario.instrucciones_recepcion = data.get('instrucciones', formulario.instrucciones_recepcion)
            formulario.eta_fecha = data.get('etaDate') or formulario.eta_fecha
            formulario.eta_hora = data.get('etaTime') or formulario.eta_hora
            
            formulario.estado = 'MODIFICADO'
            formulario.revisado_por = None
            formulario.revisado_en = None

            formulario.save()
            
            return JsonResponse({'status': 'exito', 'message': 'Formulario actualizado correctamente.'}, status=200)

        except FormularioAtencion.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Formulario no encontrado.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# |========================================================================|
# |             VISTAS DE API (Recepción y Ambulancia)                     |
# |========================================================================|    

# Vista para que Recepción obtenga formularios pendientes
@login_required
def obtener_formularios_pendientes(request):
    try:
        if not (request.user.groups.filter(name='Recepcion').exists() or request.user.is_superuser):
            return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)

        limite_tiempo = timezone.now() - timezone.timedelta(minutes=5)
        formularios_bloqueados = FormularioAtencion.objects.filter(
            revisado_en__lt=limite_tiempo 
        )
        for form in formularios_bloqueados:
            form.revisado_por = None
            form.revisado_en = None
            form.save(update_fields=['revisado_por', 'revisado_en'])
            
        formularios = FormularioAtencion.objects.filter(
            estado__in=['PENDIENTE', 'REQUIERE_MODIFICACION', 'MODIFICADO']
        ).select_related('revisado_por', 'creado_por').order_by('-creado_en') 

        lista_formularios = []
        for form in formularios:
            lista_formularios.append({
                'id': str(form.id),
                'nombre_paciente': form.nombre_paciente or "N/N",
                'estado': form.estado, 
                'estado_display': form.get_estado_display(),
                'triage': form.triage,
                'creado_por_nombre': form.creado_por.get_full_name() if form.creado_por else "Desconocido",
                'creado_en': timezone.localtime(form.creado_en).strftime('%H:%M %p'),
                'revisado_por_nombre': form.revisado_por.get_full_name() if form.revisado_por else None
            })
        
        return JsonResponse({ 'status': 'exito', 'formularios': lista_formularios })
    except Exception as e:
        return JsonResponse({ 'status': 'error', 'message': str(e) }, status=500)

# Vista para que Recepción obtenga formularios aprobados
@login_required
def obtener_formularios_aprobados(request):
    if not (request.user.groups.filter(name='Recepcion').exists() or request.user.is_superuser):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)
    formularios = FormularioAtencion.objects.filter(estado='APROBADO').order_by('-creado_en')
    lista_formularios = []
    for form in formularios:
        lista_formularios.append({
            'id': str(form.id),
            'nombre_paciente': form.nombre_paciente or "N/N",
            'triage': form.triage,
            'creado_por_nombre': form.creado_por.get_full_name() if form.creado_por else "Desconocido",
            'creado_en': timezone.localtime(form.creado_en).strftime('%H:%M %p'),
            'aprobado_por_nombre': form.aprobado_por.get_full_name() if form.aprobado_por else "N/A"
        })
    return JsonResponse({'status': 'exito', 'formularios': lista_formularios})

# Vista para que Ambulancia obtenga sus propios formularios
@login_required
def obtener_formularios_ambulancia(request):
    if not request.user.groups.filter(name='Ambulancia').exists():
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)

    try:
        limite_tiempo = timezone.now() - timezone.timedelta(minutes=5)
        formularios_bloqueados = FormularioAtencion.objects.filter(
            creado_por=request.user, 
            revisado_en__lt=limite_tiempo 
        )
        for form in formularios_bloqueados:
            form.revisado_por = None
            form.revisado_en = None
            form.save(update_fields=['revisado_por', 'revisado_en'])

        formularios = FormularioAtencion.objects.filter(
            creado_por=request.user
        ).select_related('revisado_por').order_by('-creado_en') 

        lista_formularios = []
        for form in formularios:
            lista_formularios.append({
                'id': str(form.id),
                'nombre_paciente': form.nombre_paciente or "N/A",
                'estado': form.estado, 
                'estado_display': form.get_estado_display(),
                'creado_en': form.creado_en.strftime("%d %b %Y, %H:%M"), 
                'revisado_por_nombre': form.revisado_por.get_full_name() if form.revisado_por else None,
                'chat_url': reverse('chat_formulario', args=[form.id]),
                'edit_url': reverse('editar_formulario', args=[form.id]),
            })
        
        return JsonResponse({ 'status': 'exito', 'formularios': lista_formularios })
    
    except Exception as e:
        return JsonResponse({ 'status': 'error', 'message': str(e) }, status=500)

# Vista para que Recepción obtenga detalles de un formulario
@login_required
def obtener_detalle_formulario(request, form_id):
    if not (request.user.groups.filter(name='Recepcion').exists() or request.user.is_superuser):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)
    try:
        form = FormularioAtencion.objects.get(id=form_id)
        
        detalles = { 
            'id': str(form.id), 
            'estado': form.estado, 
            'triage': form.triage,
            
            # --- Paciente ---
            'nombre_paciente': form.nombre_paciente or "N/N", 
            'rut_paciente': form.rut_paciente or "S/R", 
            'edad_paciente': form.edad_paciente, 
            'unidad_edad': form.unidad_edad_paciente, 
            'sexo': form.sexo_paciente or "No esp.", 
            'prevision': form.prevision or "No esp.",
            'accidente_laboral': "Sí" if form.accidente_laboral else "No",
            
            # --- Signos Vitales ---
            'presion_arterial': form.presion_arterial or "--/--", 
            'frecuencia_cardiaca': form.frecuencia_cardiaca or "--", 
            'frecuencia_respiratoria': form.frecuencia_respiratoria or "--",
            'temperatura': form.temperatura or "--",
            'saturacion_oxigeno': form.saturacion_oxigeno or "--", 
            'fio2': form.fio2 or "--",
            'glasgow': form.glasgow or "--",
            'llene_capilar': form.llene_capilar or "No esp.",
            'score_mottling': form.score_mottling or "No esp.",
            'musculatura_accesoria': "Sí" if form.musculatura_accesoria else "No",

            # --- Anamnesis ---
            'motivo_consulta': form.motivo_consulta or "No esp.", 
            'antecedentes': form.antecedentes or "No esp.", 
            'prestacion_requerida': form.prestacion_requerida or "No esp.",
            'funcionalidad': form.funcionalidad or "No esp.",
            'tratamiento_administrado': form.tratamiento_administrado or "No esp.",
            'solicitudes_paciente': form.solicitudes_paciente or "No esp.",

            # --- Notificación ---
            'instrucciones_recepcion': form.instrucciones_recepcion or "No esp.",
            'eta_fecha': form.eta_fecha.strftime('%Y-%m-%d') if form.eta_fecha else None,
            'eta_hora': form.eta_hora.strftime('%H:%M') if form.eta_hora else None,

            # --- Personal ---
            'creado_por_nombre': form.creado_por.get_full_name() if form.creado_por else "Desconocido", 
            'creado_en_fecha': timezone.localtime(form.creado_en).strftime('%d/%m/%Y %H:%M'), 
        }
        
        return JsonResponse({'status': 'exito', 'detalles': detalles})
    except FormularioAtencion.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Formulario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# Vista para que Recepción apruebe un formulario
@login_required
def aprobar_formulario_view(request):
    if not (request.user.groups.filter(name='Recepcion').exists() or request.user.is_superuser):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)
    if request.method == 'POST':
        try:
            form_id = request.POST.get('form_id')
            formulario = FormularioAtencion.objects.get(id=form_id)
            
            formulario.estado = 'APROBADO'
            formulario.aprobado_por = request.user 
            formulario.aprobado_en = timezone.now() # Guardar la fecha de aprobación
            formulario.revisado_por = None 
            formulario.revisado_en = None
            
            formulario.save(update_fields=['estado', 'aprobado_por', 'aprobado_en', 'revisado_por', 'revisado_en'])
            
            return JsonResponse({'status': 'exito', 'message': 'Formulario aprobado'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# Vista para que Recepción solicite modificaciones a un formulario
@login_required
def solicitar_modificacion_view(request):
    if not (request.user.groups.filter(name='Recepcion').exists() or request.user.is_superuser):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)
    if request.method == 'POST':
        try:
            form_id = request.POST.get('form_id')
            formulario = FormularioAtencion.objects.get(id=form_id)
            
            formulario.estado = 'REQUIERE_MODIFICACION' 
            formulario.aprobado_por = None 
            formulario.aprobado_en = None # Limpiar fecha de aprobación si se rechaza
            formulario.revisado_por = None 
            formulario.revisado_en = None
            
            formulario.save(update_fields=['estado', 'aprobado_por', 'aprobado_en', 'revisado_por', 'revisado_en'])
            
            return JsonResponse({'status': 'exito', 'message': 'Modificación solicitada'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# |========================================================================|
# |           VISTAS DE ELIMINACIÓN Y BLOQUEO DE FORMULARIOS               |
# |========================================================================|

# Vista para eliminar un formulario (solo Superusuario)
@login_required
def eliminar_formulario_view(request, form_id):
    if not (request.user.is_superuser or request.user.groups.filter(name='Supervisor').exists()):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)
    
    if request.method == 'POST':
        try:
            formulario = FormularioAtencion.objects.get(id=form_id)
            form_id_str = str(formulario.id) # Guardar el ID para el mensaje
            
            # La acción de eliminar
            formulario.delete() 
            
            return JsonResponse({'status': 'exito', 'message': f'Formulario {form_id_str} eliminado permanentemente'})
        
        except FormularioAtencion.DoesNotExist:
             return JsonResponse({'status': 'error', 'message': 'Formulario no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    # Si no es POST
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# Vista para bloquear un formulario
@login_required
def bloquear_formulario_view(request, form_id):

    if not (request.user.groups.filter(name='Recepcion').exists() or request.user.is_superuser):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)

    if request.method == 'POST':
        try:
            formulario = FormularioAtencion.objects.get(id=form_id)
            ahora = timezone.now()
            limite_tiempo = ahora - timezone.timedelta(minutes=5)

            if (formulario.revisado_por is not None and
                formulario.revisado_por != request.user and 
                formulario.revisado_en is not None and
                formulario.revisado_en > limite_tiempo):
                
                return JsonResponse({
                    'status': 'error_bloqueado', 
                    'message': f"El formulario está siendo revisado por: {formulario.revisado_por.get_full_name() or formulario.revisado_por.username}"
                }, status=409) 

            formulario.revisado_por = request.user
            formulario.revisado_en = ahora
            formulario.save(update_fields=['revisado_por', 'revisado_en'])
            
            return JsonResponse({'status': 'exito', 'message': 'Formulario bloqueado para revisión'})

        except FormularioAtencion.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Formulario no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# Vista para liberar un formulario
@login_required
def liberar_formulario_view(request, form_id):
    if not (request.user.groups.filter(name='Recepcion').exists() or request.user.is_superuser):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)

    if request.method == 'POST':
        try:
            formulario = FormularioAtencion.objects.get(id=form_id)
            
            if formulario.revisado_por == request.user:
                formulario.revisado_por = None
                formulario.revisado_en = None
                formulario.save(update_fields=['revisado_por', 'revisado_en'])
                return JsonResponse({'status': 'exito', 'message': 'Formulario liberado'})
            
            return JsonResponse({'status': 'sin_cambios', 'message': 'No se requiere liberación'})

        except FormularioAtencion.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Formulario no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# |========================================================================|
# |             VISTAS DE CHAT, AUDITORÍA Y EXPORTACIÓN                    |
# |========================================================================|

# Vista del chat asociado a un formulario
@login_required
def chat_formulario_view(request, form_id):
    try:
        formulario = FormularioAtencion.objects.get(id=form_id)
        es_creador = (formulario.creado_por == request.user)
        es_recepcion = request.user.groups.filter(name='Recepcion').exists()
        es_supervisor = (request.user.is_superuser or request.user.groups.filter(name='Supervisor').exists())
        if not (es_creador or es_recepcion or es_supervisor):
            messages.error(request, 'No tienes permiso para ver este chat.')
            return redirect('menu_ambulancia') 
        context = { 'formulario': formulario, 'form_id_str': str(form_id) }
        return render(request, 'chat_formulario.html', context)
    except FormularioAtencion.DoesNotExist:
        messages.error(request, 'El formulario no existe.')
        return redirect('menu_ambulancia') 

# Vista para obtener mensajes del chat
@login_required
def obtener_mensajes_chat(request, form_id):
    try:
        formulario = FormularioAtencion.objects.get(id=form_id)
        # Se usa prefetch_related para la relación ManyToMany con Groups
        mensajes = MensajeChat.objects.filter(formulario=formulario).select_related('autor').prefetch_related('autor__groups') 
        lista_mensajes = []
        for msg in mensajes:
            autor_nombre = "Usuario Eliminado"
            autor_rol = "N/A"
            if msg.autor:
                autor_nombre = msg.autor.get_full_name() if msg.autor.get_full_name() else msg.autor.username
                autor_rol = msg.autor.groups.first().name if msg.autor and msg.autor.groups.exists() else "N/A"
            
            lista_mensajes.append({
                'autor_nombre': autor_nombre,
                'autor_rol': autor_rol,
                'mensaje': msg.mensaje,
                'timestamp': timezone.localtime(msg.timestamp).strftime('%H:%M %p') 
            })
        return JsonResponse({'status': 'exito', 'mensajes': lista_mensajes})
    except FormularioAtencion.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Formulario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# Vista para enviar un mensaje al chat
@login_required
def enviar_mensaje_chat(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            form_id = data.get('form_id')
            mensaje_texto = data.get('message')
            if not form_id or not mensaje_texto:
                return JsonResponse({'status': 'error', 'message': 'Faltan datos'}, status=400)
            formulario = FormularioAtencion.objects.get(id=form_id)
            nuevo_mensaje = MensajeChat( formulario=formulario, autor=request.user, mensaje=mensaje_texto )
            nuevo_mensaje.save()
            mensaje_formateado = {
                'autor_nombre': request.user.get_full_name(),
                'autor_rol': request.user.groups.first().name if request.user.groups.exists() else "N/A",
                'mensaje': nuevo_mensaje.mensaje,
                'timestamp': timezone.localtime(nuevo_mensaje.timestamp).strftime('%H:%M %p')
            }
            return JsonResponse({ 'status': 'exito', 'nuevo_mensaje': mensaje_formateado }, status=201)
        except FormularioAtencion.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Formulario no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# Vista para obtener registros de auditoría de un usuario
@login_required
def obtener_registros_auditoria(request, user_id):
    User = get_user_model() 
    
    if not (request.user.is_superuser or request.user.groups.filter(name='Supervisor').exists()):
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)
        
    try:
        usuario_auditado = User.objects.get(id=user_id) 
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

    # 1. Acciones (Logs) - SE MANTIENE PARA TODOS
    logs_acciones = LogEntry.objects.filter(user_id=usuario_auditado.id).order_by('-action_time')[:20]
    acciones_data = []
    for log in logs_acciones:
        acciones_data.append({ 'timestamp': log.action_time.strftime('%Y-%m-%d %H:%M:%S'), 'type': log.get_change_message(), 'detail': str(log.object_repr) })

    # 2. Comunicaciones (Chat) - SE MANTIENE PARA SUPERVISOR/ADMIN
    logs_chat = MensajeChat.objects.filter( autor=usuario_auditado ).select_related('autor').prefetch_related('autor__groups').order_by('-timestamp')[:20] 
    chat_data = []
    for msg in logs_chat:
        autor_nombre = "Usuario Eliminado"
        autor_rol = "N/A"
        
        if msg.autor:
            autor_nombre = msg.autor.get_full_name() if msg.autor.get_full_name() else msg.autor.username
            autor_rol = msg.autor.groups.first().name if msg.autor.groups.exists() else 'N/A'
        
        chat_data.append({ 
            'autor_nombre': autor_nombre,
            'autor_rol': autor_rol,
            'mensaje': msg.mensaje, 
            'timestamp': msg.timestamp.strftime('%I:%M %p') 
        })

    # 3. Formularios Creados (para Ambulancia)
    formularios = FormularioAtencion.objects.filter(creado_por=usuario_auditado).order_by('-creado_en')
    formularios_data = []
    for form in formularios:
        formularios_data.append({
            'id': str(form.id),
            'nombre_paciente': form.nombre_paciente or "N/N",
            'creado_en': form.creado_en.strftime("%d-%m-%Y"),
            'estado': form.estado,
            'estado_display': form.get_estado_display(),
        })
        
    # 4. Formularios Gestionados (para Recepción)
    formularios_gestionados_data = []
    
    form_ids_from_chat = MensajeChat.objects.filter(autor=usuario_auditado).values_list('formulario_id', flat=True).distinct()

    formularios_gestionados = FormularioAtencion.objects.filter(
        Q(aprobado_por=usuario_auditado) |
        Q(id__in=list(form_ids_from_chat))
    ).distinct().order_by('-creado_en')
    
    for form in formularios_gestionados:
        formularios_gestionados_data.append({
            'id': str(form.id),
            'nombre_paciente': form.nombre_paciente or "N/N",
            'creado_en': form.creado_en.strftime("%d-%m-%Y"),
            'estado': form.estado,
            'estado_display': form.get_estado_display(),
        })

    return JsonResponse({ 
        'user_name': usuario_auditado.get_full_name() or usuario_auditado.username, 
        'user_role': usuario_auditado.groups.first().name if usuario_auditado.groups.exists() else 'N/A', 
        'actions': acciones_data, 
        'chat': chat_data,
        'formularios': formularios_data, 
        'formularios_gestionados': formularios_gestionados_data,
    })

@login_required
def api_obtener_todos_los_formularios(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='Supervisor').exists()):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)
        
    try:
        formularios = FormularioAtencion.objects.select_related('creado_por').all().order_by('-creado_en')
        lista_formularios = []
        for form in formularios:
            lista_formularios.append({
                'id': str(form.id),
                'nombre_paciente': form.nombre_paciente or "N/N",
                'creado_en': form.creado_en.strftime("%d-%m-%Y"),
                'aprobado_en': form.aprobado_en.strftime("%Y-%m-%d") if form.aprobado_en else None,
                'estado': form.estado,
                'estado_display': form.get_estado_display(),
                'creado_por_nombre': form.creado_por.username if form.creado_por else "Desconocido",
            })
        return JsonResponse({'status': 'exito', 'formularios': lista_formularios})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# ==================================
# API CREAR USUARIO
# ==================================
@login_required
def api_crear_usuario_view(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='Supervisor').exists()):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            User = get_user_model()

            username = data.get('username')
            password = data.get('password')
            role_name = data.get('role')
            rut = data.get('rut') 

            if not (username and password and role_name and rut):
                return JsonResponse({'status': 'error', 'message': 'Username, Contraseña, Rol y RUT son requeridos.'}, status=400)

            if User.objects.filter(username=username).exists():
                return JsonResponse({'status': 'error', 'message': f'El username "{username}" ya existe.'}, status=400)

            if User.objects.filter(rut=rut).exists():
                return JsonResponse({'status': 'error', 'message': f'El RUT "{rut}" ya está registrado a nombre de otro usuario.'}, status=400)

            user = User.objects.create_user(username=username, password=password)

            user.first_name = data.get('first_name', '')
            user.last_name = data.get('last_name', '')
            user.email = data.get('email', '')
            user.rut = rut 
            user.save()

            try:
                group = Group.objects.get(name=role_name)
                user.groups.add(group)
                
                if role_name == 'Supervisor':
                    user.is_staff = True
                    user.is_superuser = True
                    user.save()
                    
            except Group.DoesNotExist:
                user.delete()
                return JsonResponse({'status': 'error', 'message': f'El rol "{role_name}" no existe en el sistema.'}, status=400)
            
            return JsonResponse({'status': 'exito', 'message': f'Usuario "{username}" creado con el rol "{role_name}".'}, status=201)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# ==================================
# API: EDITAR USUARIO (CORREGIDA)
# ==================================
@login_required
def api_editar_usuario_view(request, user_id):
    if not (request.user.is_superuser or request.user.groups.filter(name='Supervisor').exists()):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)
    
    User = get_user_model()
    try:
        usuario_a_editar = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Usuario no encontrado'}, status=404)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Actualizar campos
            usuario_a_editar.first_name = data.get('first_name', usuario_a_editar.first_name)
            usuario_a_editar.last_name = data.get('last_name', usuario_a_editar.last_name)
            usuario_a_editar.email = data.get('email', usuario_a_editar.email)
            usuario_a_editar.rut = data.get('rut', usuario_a_editar.rut)
            
            # Actualizar contraseña (si se envió una nueva)
            if data.get('password'):
                usuario_a_editar.set_password(data.get('password'))
            
            # Actualizar Rol
            role_name = data.get('role')
            if role_name:
                group = Group.objects.get(name=role_name)
                usuario_a_editar.groups.clear() # Limpiar roles antiguos
                usuario_a_editar.groups.add(group)
                
                # Ajustar permisos de staff/superuser si es Supervisor
                if role_name == 'Supervisor':
                    usuario_a_editar.is_staff = True
                    usuario_a_editar.is_superuser = True
                else:
                    usuario_a_editar.is_staff = False
                    usuario_a_editar.is_superuser = False

            usuario_a_editar.save()
            return JsonResponse({'status': 'exito', 'message': 'Usuario actualizado'})
            
        except Group.DoesNotExist:
             return JsonResponse({'status': 'error', 'message': f'El rol "{role_name}" no existe'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# ==================================
# API: ELIMINAR USUARIO (CORREGIDA)
# ==================================
@login_required
def api_eliminar_usuario_view(request, user_id):
    if not (request.user.is_superuser or request.user.groups.filter(name='Supervisor').exists()):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)
    
    User = get_user_model()
    try:
        usuario_a_eliminar = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Usuario no encontrado'}, status=404)

    if request.method == 'POST': # Usamos POST por seguridad (con CSRF)
        try:
            username = usuario_a_eliminar.username
            usuario_a_eliminar.delete()
            return JsonResponse({'status': 'exito', 'message': f'Usuario {username} eliminado'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

# ==================================
# API: AUDITORÍA DEL SISTEMA (CORREGIDA)
# ==================================
@login_required
def api_obtener_logs_del_sistema(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='Supervisor').exists()):
        return JsonResponse({'status': 'error', 'message': 'Acceso no autorizado'}, status=403)
    
    try:
        # Obtener todos los logs, con el nombre de usuario
        logs = LogEntry.objects.all().select_related('user').order_by('-action_time')
        
        lista_logs = []
        for log in logs:
            lista_logs.append({
                'timestamp': log.action_time.strftime('%Y-%m-%d %H:%M:%S'),
                'user': log.user.username if log.user else 'Sistema',
                'type': log.get_change_message(), # 'Added.', 'Changed.', 'Deleted.'
                'detail': str(log.object_repr)
            })
        
        return JsonResponse({'status': 'exito', 'logs': lista_logs})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# ==================================
# API DE EXPORTACIÓN (MODIFICADA)
# ==================================
@login_required
def exportar_formulario_view(request, form_id, formato):
    try:
        formulario = FormularioAtencion.objects.get(id=form_id)
    except FormularioAtencion.DoesNotExist:
        return HttpResponse("Formulario no encontrado", status=404)
    if formato == 'pdf':
        filename = f"caso_{formulario.nombre_paciente or 'NN'}_{form_id[:8]}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        generar_pdf(response, formulario) # <-- Llamada a la función PDF actualizada
        return response
    elif formato == 'xlsx' or formato == 'excel': 
        filename = f"caso_{formulario.nombre_paciente or 'NN'}_{form_id[:8]}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        generar_excel_actualizado(response, formulario) # <-- Llamada a la función Excel actualizada
        return response
    else:
        return HttpResponse(f"Formato no válido. Se recibió '{formato}'", status=400)

# ================== -->
# INICIO CAMBIOS 3.4 -->
# (Función de Excel - Sin cambios esta vez, ya estaba bien)
# ================== -->
def generar_excel_actualizado(response, formulario):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Caso {formulario.id.hex[:8]}"
    
    # --- Estilos Visuales ---
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True, color="FFFFFF") 
    header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid") # Rojo
    label_font = Font(bold=True)
    label_align = Alignment(horizontal='right', vertical='top')
    value_align = Alignment(vertical='top', wrap_text=True)

    # --- Funciones de ayuda ---
    def add_row(label, value, row_num):
        ws[f'A{row_num}'] = label
        ws[f'A{row_num}'].font = label_font
        ws[f'A{row_num}'].alignment = label_align
        ws[f'B{row_num}'] = value
        ws[f'B{row_num}'].alignment = value_align
        # Ajuste dinámico de altura de fila
        if value:
            lines = str(value).split('\n')
            max_line_len = 0
            for line in lines:
                max_line_len = max(max_line_len, len(line))
            
            num_lines = len(lines)
            # Aumentar líneas basado en el ancho
            num_lines += (max_line_len // 50) * len(lines)
            ws.row_dimensions[row_num].height = max(15, num_lines * 15)
        else:
            ws.row_dimensions[row_num].height = 15
        
    def add_header(text, row_num):
        ws.merge_cells(f'A{row_num}:B{row_num}')
        cell = ws[f'A{row_num}']
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row_num].height = 20

    # --- Título ---
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Reporte de Atención - Caso #{formulario.id.hex[:8]}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25

    # --- Configuración de Columnas ---
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50

    row = 3
    
    # --- Datos del Paciente ---
    add_header("1. Datos del Paciente", row); row += 1
    add_row("Paciente:", formulario.nombre_paciente or "N/A", row); row += 1
    add_row("RUT:", formulario.rut_paciente or "N/A", row); row += 1
    add_row("Edad:", f"{formulario.edad_paciente or '?'} {formulario.unidad_edad_paciente}", row); row += 1
    add_row("Sexo:", formulario.sexo_paciente or "N/A", row); row += 1
    add_row("Previsión:", formulario.get_prevision_display() or "N/A", row); row += 1
    add_row("Accidente Laboral:", "Sí" if formulario.accidente_laboral else "No", row); row += 2 
    
    # --- Signos Vitales y Triage ---
    add_header("2. Evaluación Clínica y Signos Vitales", row); row += 1
    add_row("Triage:", formulario.triage or "N/A", row); row += 1
    add_row("Presión Arterial:", formulario.presion_arterial or "N/A", row); row += 1
    add_row("Frec. Cardiaca:", f"{formulario.frecuencia_cardiaca} lpm" if formulario.frecuencia_cardiaca else "N/A", row); row += 1
    add_row("Frec. Respiratoria:", f"{formulario.frecuencia_respiratoria} rpm" if formulario.frecuencia_respiratoria else "N/A", row); row += 1
    add_row("Temperatura:", f"{formulario.temperatura} °C" if formulario.temperatura else "N/A", row); row += 1
    add_row("Saturación O₂:", f"{formulario.saturacion_oxigeno} %" if formulario.saturacion_oxigeno else "N/A", row); row += 1
    add_row("FiO2:", f"{formulario.fio2} %" if formulario.fio2 else "N/A", row); row += 1
    add_row("Glasgow:", formulario.glasgow or "N/A", row); row += 1
    add_row("Llene Capilar:", formulario.llene_capilar or "N/A", row); row += 1
    add_row("Score Mottling:", formulario.score_mottling or "N/A", row); row += 1
    add_row("Uso Musculatura:", "Sí" if formulario.musculatura_accesoria else "No", row); row += 2
    
    # --- Anamnesis y Notificación ---
    add_header("3. Anamnesis y Motivo de Traslado", row); row += 1
    add_row("Motivo Consulta:", formulario.motivo_consulta or "N/A", row); row += 1
    add_row("Prestación Requerida:", formulario.prestacion_requerida or "N/A", row); row += 1
    add_row("Antecedentes:", formulario.antecedentes or "N/A", row); row += 1
    add_row("Funcionalidad:", formulario.funcionalidad or "N/A", row); row += 1
    add_row("Tratamiento Admin.:", formulario.tratamiento_administrado or "N/A", row); row += 1
    add_row("Solicitudes Paciente:", formulario.solicitudes_paciente or "N/A", row); row += 2
    
    # --- Notificación ---
    add_header("4. Notificación", row); row += 1
    add_row("Instrucciones Recepción:", formulario.instrucciones_recepcion or "N/A", row); row += 1
    eta_fecha_str = formulario.eta_fecha.strftime('%d-%m-%Y') if formulario.eta_fecha else "N/A"
    eta_hora_str = formulario.eta_hora.strftime('%H:%M') if formulario.eta_hora else "N/A"
    add_row("ETA:", f"{eta_fecha_str} a las {eta_hora_str}", row); row += 2

    # --- Personal ---
    add_header("5. Personal", row); row += 1
    add_row("Personal Ambulancia:", formulario.creado_por.get_full_name() if formulario.creado_por else "N/A", row); row += 1
    add_row("Fecha Creación:", timezone.localtime(formulario.creado_en).strftime('%d-%m-%Y %H:%M'), row); row += 1
    if formulario.aprobado_por:
        add_row("Aprobado Por:", formulario.aprobado_por.get_full_name(), row); row += 1
        add_row("Fecha Aprobación:", timezone.localtime(formulario.aprobado_en).strftime('%d-%m-%Y %H:%M'), row); row += 1
    
    wb.save(response)

# ================== -->
# INICIO CAMBIOS 4.2 -->
# (Función de PDF corregida y robustecida)
# ================== -->
def generar_pdf(response, formulario):
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter 
    margin = 0.75 * inch
    
    class PDFCursor:
        def __init__(self, x, y_start):
            self.x = x
            self.y_start = y_start
            self.y = y_start
            self.line_height = 16 
            self.field_indent = 1.5 * inch

        def check_page_break(self, needed_space=40):
            # Comprueba si se necesita espacio, si no, salta de página
            if self.y < margin + needed_space: 
                p.showPage()
                self.y = self.y_start
                
        def write_header(self, text):
            # Escribe un cabecera de sección
            self.check_page_break(60) # Espacio para cabecera + 1 campo
            p.setFont("Helvetica-Bold", 12)
            p.setFillColorRGB(0.93, 0.26, 0.26) # Rojo (EF4444)
            p.drawString(self.x, self.y + 10, text)
            self.y -= self.line_height * 0.5
            p.line(self.x, self.y + 8, width - margin, self.y + 8)
            self.y -= self.line_height * 1.5
            p.setFillColorRGB(0, 0, 0) # Reset color a negro

        def write_field(self, label, value):
            # Escribe un campo simple (Etiqueta: Valor)
            self.check_page_break(20)
            p.setFont("Helvetica-Bold", 10)
            p.drawString(self.x, self.y, label)
            p.setFont("Helvetica", 10)
            p.drawString(self.x + self.field_indent, self.y, str(value or "N/A")) 
            self.y -= self.line_height * 1.5 

        def write_multiline_field(self, label, value):
            # Escribe un campo largo, dividiendo el texto en múltiples líneas
            self.check_page_break(60) # Mínimo 60 de espacio
            p.setFont("Helvetica-Bold", 10)
            p.drawString(self.x, self.y, label)
            p.setFont("Helvetica", 10)

            text_obj = p.beginText(self.x + self.field_indent, self.y)
            
            val = str(value or "N/A")
            max_width_chars = 75 # Caracteres aproximados por línea
            
            lines_drawn = 0
            # Divide por saltos de línea manuales primero
            for line in val.split('\n'):
                # Envuelve líneas largas
                wrapped_lines = textwrap.wrap(line, width=max_width_chars)
                if not wrapped_lines: # Si es una línea vacía
                    text_obj.textLine(" ") 
                    lines_drawn += 1
                for wrapped_line in wrapped_lines:
                    text_obj.textLine(wrapped_line)
                    lines_drawn += 1
            
            p.drawText(text_obj)
            
            # Mover el cursor 'y' hacia abajo basado en líneas dibujadas
            self.y -= (lines_drawn * 12) # 12 es la altura de línea
            self.y -= self.line_height * 0.5 # Espacio extra post-campo
    
    c = PDFCursor(margin, height - margin)
    
    # --- Título ---
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width / 2.0, c.y, f"Reporte de Atención - Caso #{formulario.id.hex[:8]}")
    c.y -= 40 
    
    # --- Datos del Paciente ---
    c.write_header("1. Datos del Paciente")
    c.write_field("Paciente:", formulario.nombre_paciente)
    c.write_field("RUT:", formulario.rut_paciente)
    c.write_field("Edad:", f"{formulario.edad_paciente or '?'} {formulario.unidad_edad_paciente}")
    c.write_field("Sexo:", formulario.sexo_paciente)
    c.write_field("Previsión:", formulario.get_prevision_display() or "N/A")
    c.write_field("Accidente Laboral:", "Sí" if formulario.accidente_laboral else "No")
    
    # --- Evaluación Clínica ---
    c.write_header("2. Evaluación Clínica y Signos Vitales")
    c.write_field("Triage:", formulario.triage)
    c.write_field("Presión Arterial:", formulario.presion_arterial)
    c.write_field("Frec. Cardiaca:", f"{formulario.frecuencia_cardiaca} lpm" if formulario.frecuencia_cardiaca else "N/A")
    c.write_field("Frec. Respiratoria:", f"{formulario.frecuencia_respiratoria} rpm" if formulario.frecuencia_respiratoria else "N/A")
    c.write_field("Temperatura:", f"{formulario.temperatura} °C" if formulario.temperatura else "N/A")
    c.write_field("Saturación O₂:", f"{formulario.saturacion_oxigeno} %" if formulario.saturacion_oxigeno else "N/A")
    c.write_field("FiO2:", f"{formulario.fio2} %" if formulario.fio2 else "N/A")
    c.write_field("Glasgow:", formulario.glasgow)
    c.write_field("Llene Capilar:", formulario.llene_capilar)
    c.write_field("Score Mottling:", formulario.score_mottling)
    c.write_field("Uso Musculatura:", "Sí" if formulario.musculatura_accesoria else "No")
    
    c.check_page_break(200) # Comprobar si hay espacio para la siguiente sección
    
    # --- Anamnesis ---
    c.write_header("3. Anamnesis y Motivo de Traslado")
    c.write_multiline_field("Motivo Consulta:", formulario.motivo_consulta)
    c.write_multiline_field("Prestación Requerida:", formulario.prestacion_requerida)
    c.write_multiline_field("Antecedentes:", formulario.antecedentes)
    c.write_field("Funcionalidad:", formulario.funcionalidad)
    c.write_multiline_field("Tratamiento Admin.:", formulario.tratamiento_administrado)
    c.write_multiline_field("Solicitudes Paciente:", formulario.solicitudes_paciente)
    
    c.check_page_break(150) # Comprobar si hay espacio para la siguiente sección

    # --- Notificación ---
    c.write_header("4. Notificación")
    c.write_multiline_field("Instrucciones Recepción:", formulario.instrucciones_recepcion) 
    eta_fecha_str = formulario.eta_fecha.strftime('%d-%m-%Y') if formulario.eta_fecha else "N/A"
    eta_hora_str = formulario.eta_hora.strftime('%H:%M') if formulario.eta_hora else "N/A"
    c.write_field("ETA:", f"{eta_fecha_str} a las {eta_hora_str}")
    
    # --- Personal ---
    c.write_header("5. Personal")
    c.write_field("Personal Ambulancia:", formulario.creado_por.get_full_name() if formulario.creado_por else "N/A")
    c.write_field("Fecha Creación:", timezone.localtime(formulario.creado_en).strftime('%d-%m-%Y %H:%M'))
    if formulario.aprobado_por:
        c.write_field("Aprobado Por:", formulario.aprobado_por.get_full_name())
        c.write_field("Fecha Aprobación:", timezone.localtime(formulario.aprobado_en).strftime('%d-%m-%Y %H:%M'))
    
    p.showPage()
    p.save()
# ================== -->
# FIN CAMBIOS 4.2    -->
# ================== -->